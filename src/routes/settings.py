"""账号设置路由模块"""
from io import BytesIO
from datetime import datetime

from flask import Blueprint, render_template, request, session, redirect, url_for, flash, jsonify, send_file
from openpyxl import Workbook

from models import (db, User, Transaction, AccountBalance, BabyFund,
                    SavingsRecord, SavingsPlan, Account, AccountType, Category)

settings_bp = Blueprint('settings', __name__, url_prefix='/settings')


@settings_bp.route('/')
def settings_page():
    """账号设置页"""
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('auth.login'))

    user = User.query.get(user_id)
    family = user.family if user else None

    return render_template('settings.html',
                           user=user,
                           family=family,
                           username=session.get('nickname', session.get('username', '用户')),
                           page_title='账号设置')


@settings_bp.route('/nickname', methods=['POST'])
def update_nickname():
    """修改昵称"""
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'error': '未登录'}), 401

    user = User.query.get(user_id)
    nickname = request.form.get('nickname', '').strip()

    if not nickname or len(nickname) > 20:
        flash('昵称不能为空且不超过20字', 'error')
        return redirect(url_for('settings.settings_page'))

    user.nickname = nickname
    session['nickname'] = nickname  # 同步 session
    db.session.commit()

    flash('昵称已更新', 'success')
    return redirect(url_for('settings.settings_page'))


@settings_bp.route('/avatar', methods=['POST'])
def update_avatar():
    """修改展示图标"""
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'error': '未登录'}), 401

    user = User.query.get(user_id)
    avatar_text = request.form.get('avatar_text', '').strip()

    if not avatar_text or len(avatar_text) > 2:
        flash('图标需为1-2个字符', 'error')
        return redirect(url_for('settings.settings_page'))

    user.avatar_text = avatar_text
    db.session.commit()

    flash('展示图标已更新', 'success')
    return redirect(url_for('settings.settings_page'))


@settings_bp.route('/password', methods=['POST'])
def update_password():
    """修改密码"""
    user_id = session.get('user_id')
    if not user_id:
        return jsonify({'success': False, 'error': '未登录'}), 401

    user = User.query.get(user_id)
    old_password = request.form.get('old_password', '')
    new_password = request.form.get('new_password', '')
    confirm_password = request.form.get('confirm_password', '')

    # 验证旧密码
    if not user.check_password(old_password):
        flash('当前密码错误', 'error')
        return redirect(url_for('settings.settings_page'))

    # 验证新密码规则
    if len(new_password) < 8:
        flash('新密码至少8位', 'error')
        return redirect(url_for('settings.settings_page'))

    has_letter = any(c.isalpha() for c in new_password)
    has_digit = any(c.isdigit() for c in new_password)
    if not (has_letter and has_digit):
        flash('新密码需包含字母和数字', 'error')
        return redirect(url_for('settings.settings_page'))

    # 验证两次输入一致
    if new_password != confirm_password:
        flash('两次输入的密码不一致', 'error')
        return redirect(url_for('settings.settings_page'))

    user.set_password(new_password)
    db.session.commit()

    flash('密码已修改', 'success')
    return redirect(url_for('settings.settings_page'))


@settings_bp.route('/export')
def export_data():
    """导出数据为 Excel 文件"""
    user_id = session.get('user_id')
    if not user_id:
        return redirect(url_for('auth.login'))

    user = User.query.get(user_id)
    scope = request.args.get('scope', 'personal')

    # 确定要导出的用户 ID 列表
    if scope == 'family' and user.family:
        user_ids = [m.id for m in user.family.members]
    else:
        user_ids = [user_id]
        scope = 'personal'  # 无家庭时强制个人

    # 构建用户名映射
    users = {u.id: u.nickname or u.username for u in User.query.filter(User.id.in_(user_ids)).all()}

    wb = Workbook()

    # --- Sheet 1: 交易记录 ---
    ws1 = wb.active
    ws1.title = '交易记录'
    ws1.append(['日期', '类型', '金额', '分类', '账户', '备注', '记录人', '创建时间'])

    type_map = {'income': '收入', 'expense': '支出', 'transfer_in': '转入', 'transfer_out': '转出'}
    transactions = Transaction.query.filter(Transaction.user_id.in_(user_ids)) \
        .order_by(Transaction.transaction_date.desc()).all()
    for t in transactions:
        ws1.append([
            t.transaction_date.strftime('%Y-%m-%d') if t.transaction_date else '',
            type_map.get(t.type, t.type),
            float(t.amount),
            t.category.name if t.category else '',
            t.account.name if t.account else '',
            t.description or '',
            users.get(t.user_id, ''),
            t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
        ])

    # --- Sheet 2: 资产快照 ---
    ws2 = wb.create_sheet('资产快照')
    ws2.append(['时间', '账户', '类型', '变更金额', '余额', '来源', '操作人'])

    source_map = {'snapshot': '快照', 'transfer': '转账'}
    # 先获取涉及的账户 ID
    account_ids = [a.id for a in Account.query.filter(Account.user_id.in_(user_ids)).all()]
    balances = AccountBalance.query.filter(AccountBalance.account_id.in_(account_ids)) \
        .order_by(AccountBalance.record_month.desc()).all() if account_ids else []
    for b in balances:
        acct = b.account
        ws2.append([
            b.record_month.strftime('%Y-%m-%d') if b.record_month else '',
            acct.name if acct else '',
            acct.account_type.name if acct and acct.account_type else '',
            float(b.change_amount) if b.change_amount else 0,
            float(b.balance),
            source_map.get(b.source, b.source or ''),
            users.get(b.recorded_by, '') if b.recorded_by else '',
        ])

    # --- Sheet 3: 宝宝基金 ---
    ws3 = wb.create_sheet('宝宝基金')
    ws3.append(['赠送日期', '赠送人', '事件类型', '金额', '备注', '记录时间'])

    baby_funds = BabyFund.query.filter(BabyFund.created_by.in_(user_ids)) \
        .order_by(BabyFund.event_date.desc()).all()
    for bf in baby_funds:
        ws3.append([
            bf.event_date.strftime('%Y-%m-%d') if bf.event_date else '',
            bf.giver_name,
            bf.event_type or '',
            float(bf.amount),
            bf.notes or '',
            bf.created_at.strftime('%Y-%m-%d %H:%M') if bf.created_at else '',
        ])

    # --- Sheet 4: 储蓄记录 ---
    ws4 = wb.create_sheet('储蓄记录')
    ws4.append(['日期', '计划名称', '金额', '备注', '记录人'])

    savings = SavingsRecord.query.filter(SavingsRecord.user_id.in_(user_ids)) \
        .order_by(SavingsRecord.record_date.desc()).all()
    for s in savings:
        ws4.append([
            s.record_date.strftime('%Y-%m-%d') if s.record_date else '',
            s.plan.name if s.plan else '',
            float(s.amount),
            s.description or '',
            users.get(s.user_id, ''),
        ])

    # 写入内存并返回
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    scope_label = '个人' if scope == 'personal' else '家庭'
    date_str = datetime.now().strftime('%Y%m%d')
    filename = f'家庭财务_{scope_label}_{date_str}.xlsx'

    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)
